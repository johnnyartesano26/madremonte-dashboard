#!/usr/bin/env python3
"""
update_bar_dashboard.py
Descarga "Nueva Sede Arqueo Bar" desde Google Sheets como XLSX,
parsea todas las pestañas y regenera el JSON de window.BAR_DATA en bar.html.
Uso: python3 update_bar_dashboard.py
"""

import json
import os
import re
from datetime import datetime
import requests
import openpyxl

# ── Config ──
SHEET_ID = "1pYhvrTfPAVzxW8qeaR8LU8pl6YhEUP81"
XLSX_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
DASHBOARD_DIR = os.path.dirname(os.path.abspath(__file__))
HTML_PATH = os.path.join(DASHBOARD_DIR, "bar.html")

# ── Mapeo nombres de mes español → número ──
MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}


def parse_sheet_name(year_str, month_str):
    """Extrae año y número de mes del nombre de la pestaña.
    Ej: 'Noviembre\\'25' -> (11, 2025), 'Enero\\'26' -> (1, 2026)"""
    year = 2000 + int(re.sub(r"[^0-9]", "", year_str))
    month = MESES.get(month_str.lower())
    if month is None:
        # Fallback: buscar con --os finales
        for k, v in MESES.items():
            if month_str.lower().startswith(k):
                month = v
                break
    return month, year


def fix_date(dt, expected_month, expected_year):
    """Corrige fechas corruptas de la hoja de Noviembre 2025."""
    if not isinstance(dt, datetime):
        return dt
    # Patrón corrupto: año entre 2001-2031, mes=expected, día=1
    if dt.year > 2000 and dt.year < 2032 and dt.month == expected_month and dt.day == 1:
        real_day = dt.year - 2000
        return datetime(expected_year, expected_month, real_day)
    return dt


def clean_money(val):
    """Convierte valor numérico a entero. Maneja None, strings y floats."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    if isinstance(val, str):
        cleaned = val.replace("$", "").replace(".", "").replace(",", ".").strip()
        if cleaned == "" or cleaned == "-":
            return 0
        try:
            return int(float(cleaned))
        except ValueError:
            return 0
    return 0


COLUMNAS_J = [
    {"key": "fecha", "label": "Fecha"},
    {"key": "cierreAlegra", "label": "Cierre Alegra", "tipo": "dinero"},
    {"key": "cierreFormato", "label": "Cierre Formato", "tipo": "dinero"},
    {"key": "efectivo", "label": "Efectivo", "tipo": "dinero"},
    {"key": "transferencias", "label": "Transferencias", "tipo": "dinero"},
    {"key": "datafono", "label": "Datafono", "tipo": "dinero"},
    {"key": "propinas", "label": "Propinas", "tipo": "dinero"},
    {"key": "observaciones", "label": "Observaciones", "tipo": "texto"},
]


def parse_sheet(ws, sheet_tab_name):
    """Parsea una pestaña y devuelve el objeto hoja para BAR_DATA."""
    # Obtener nombre real desde la celda A1
    raw_data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=9, values_only=True):
        raw_data.append(row)

    title_row = raw_data[0]
    nombre = str(title_row[0]).strip() if title_row[0] else sheet_tab_name

    # Extraer mes/año del nombre de pestaña (ej: 'Noviembre´25')
    m = re.match(r"([a-záéíóúñ]+).*?(\d+)", sheet_tab_name, re.IGNORECASE)
    if m:
        month, year = parse_sheet_name(m.group(2), m.group(1))
    else:
        month, year = 1, 2025

    # Filtrar filas de datos (saltar fila 1=title, fila 2=headers, filas con TOTAL o vacías)
    datos = []
    for row in raw_data[2:]:
        fecha_val = row[0] if len(row) > 0 else None

        # Saltar filas de TOTAL o vacías
        if fecha_val is None:
            continue
        if isinstance(fecha_val, str) and "total" in fecha_val.lower():
            continue

        # Corregir fecha
        fecha_corrected = fix_date(fecha_val, month, year) if isinstance(fecha_val, datetime) else fecha_val

        # Si no es fecha datetime, saltar
        if not isinstance(fecha_corrected, datetime):
            continue

        fecha_str = fecha_corrected.strftime("%Y-%m-%d 00:00:00")

        # Extraer valores (columnas B-H, índice 1-6; I=observaciones índice 8)
        cierreAlegra = clean_money(row[1]) if len(row) > 1 else 0
        cierreFormato = clean_money(row[2]) if len(row) > 2 else 0
        efectivo = clean_money(row[3]) if len(row) > 3 else 0
        transferencias = clean_money(row[4]) if len(row) > 4 else 0
        datafono = clean_money(row[5]) if len(row) > 5 else 0
        propinas = clean_money(row[6]) if len(row) > 6 else 0
        observaciones = str(row[8]).strip() if len(row) > 8 and row[8] else ""

        datos.append({
            "fecha": fecha_str,
            "cierreAlegra": cierreAlegra,
            "cierreFormato": cierreFormato,
            "efectivo": efectivo,
            "transferencias": transferencias,
            "datafono": datafono,
            "propinas": propinas,
            "observaciones": observaciones,
        })

    return {
        "nombre": nombre,
        "columnas": COLUMNAS_J,
        "datos": datos,
    }


def main():
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Iniciando actualización del dashboard...")

    # 1. Descargar XLSX
    print("  Descargando Google Sheet...")
    resp = requests.get(XLSX_URL, allow_redirects=True, timeout=30)
    if resp.status_code != 200:
        print(f"  ERROR: status {resp.status_code}")
        return 1

    xlsx_path = "/tmp/arqueo_bar_dashboard.xlsx"
    with open(xlsx_path, "wb") as f:
        f.write(resp.content)

    # 2. Parsear workbook
    print(f"  Parseando {len(resp.content)} bytes...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    hojas = []
    totales = {"cierreAlegra": 0, "cierreFormato": 0, "efectivo": 0,
               "transferencias": 0, "datafono": 0, "propinas": 0}
    total_dias = 0

    for tab_name in wb.sheetnames:
        ws = wb[tab_name]
        print(f"  Procesando: {tab_name}")
        hoja = parse_sheet(ws, tab_name)
        hojas.append(hoja)

        for d in hoja["datos"]:
            totales["cierreAlegra"] += d["cierreAlegra"]
            totales["cierreFormato"] += d["cierreFormato"]
            totales["efectivo"] += d["efectivo"]
            totales["transferencias"] += d["transferencias"]
            totales["datafono"] += d["datafono"]
            totales["propinas"] += d["propinas"]
            if d["cierreAlegra"] > 0 or d["cierreFormato"] > 0:
                total_dias += 1

    bar_data = {
        "hojas": hojas,
        "totales": totales,
        "totalDias": total_dias,
    }

    bar_data_json = json.dumps(bar_data, ensure_ascii=False, indent=2)

    # 3. Leer bar.html y reemplazar la sección BAR_DATA
    with open(HTML_PATH, "r", encoding="utf-8") as f:
        html = f.read()

    # Reemplazar la sección entre los marcadores
    START_MARKER = "/*+++BAR_DATA_START+++*/"
    END_MARKER = "/*+++BAR_DATA_END+++*/"
    new_block = f"window.BAR_DATA = {bar_data_json};"

    if START_MARKER in html and END_MARKER in html:
        pattern = re.compile(
            re.escape(START_MARKER) + r".*?" + re.escape(END_MARKER),
            re.DOTALL,
        )
        replacement = f"{START_MARKER}\n{new_block}\n{END_MARKER}"
        html = pattern.sub(replacement, html)
    else:
        # Fallback: buscar window.BAR_DATA = {...};
        pattern = re.compile(r"window\.BAR_DATA\s*=\s*\{.*?\};", re.DOTALL)
        if pattern.search(html):
            html = pattern.sub(f"{START_MARKER}\n{new_block}\n{END_MARKER}", html, count=1)
        else:
            print("  ERROR: No se encontró window.BAR_DATA en bar.html")
            return 1

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"  ✅ bar.html actualizado: {len(hojas)} meses, {total_dias} días con datos")
    print(f"  Totales: Alegra=${totales['cierreAlegra']:,}, Formato=${totales['cierreFormato']:,}")

    return 0


if __name__ == "__main__":
    exit(main())
